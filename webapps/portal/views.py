# webapps/portal/views.py
from __future__ import annotations

import datetime as dt
import json
import os
from typing import Any, Dict, Optional, Tuple, List

from django.conf import settings
from django.db.models import Count
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.http import require_GET

from webapps.portal.decorators import require_node
from webapps.portal.models import PortalUsageLog
from webapps.database.db_factory import db_query_all

# 仍保留 aaa decode / oracle error 供 debug 顯示，但不再當作登入主流程
from .utils import aaadecode
from .oracle_emp import get_last_error
from .oracle_emp import get_emp_name
from .oracle_emp import get_factory_plant_by_id


# =========================================================
# Portal Home
# - 依規範：登入資訊由 IISRemoteUserBridgeMiddleware 統一注入
#   request.login_user / request.login_user_name
# - index 只負責渲染、fail-open（portal 入口不可因模板 reverse 失敗整頁爆）
# =========================================================
@require_GET
def index(request: HttpRequest) -> HttpResponse:
    """
    Portal 首頁：
    - 正常回 HTML
    - DEBUG 才回詳細文字錯誤（方便你查 template/render 問題）
    """
    aaa = (request.GET.get("aaa") or "").strip()
    remote_user = (request.META.get("REMOTE_USER") or request.META.get("AUTH_USER") or "").strip()

    ctx = {
        # debug only
        "aaa": aaa,
        "aaa_decoded": aaadecode(aaa) if aaa else "",
        "remote_user": remote_user,
        "ora_emp_error": get_last_error(),
        # middleware injected
        "login_user": getattr(request, "login_user", "") or "",
        "login_user_name": getattr(request, "login_user_name", "") or "",
        "open_notebook_url": getattr(settings, "OPEN_NOTEBOOK_PORTAL_URL", "") or "http://127.0.0.1:8502",
    }

    try:
        return render(request, "portal/index.html", ctx)
    except Exception as e:
        if getattr(settings, "DEBUG", False):
            return HttpResponse(
                "Portal template render failed.\n\n"
                f"Error: {type(e).__name__}: {e}\n\n"
                f"aaa={aaa}\n"
                f"aaa_decoded={ctx['aaa_decoded']}\n"
                f"remote_user={remote_user}\n"
                f"login_user={ctx['login_user']}\n"
                f"login_user_name={ctx['login_user_name']}\n"
                f"ora_emp_error={ctx['ora_emp_error']}\n",
                content_type="text/plain; charset=utf-8",
                status=500,
            )
        return HttpResponse("Portal error", status=500)


# =========================================================
# WhoAmI
# - 依規範：用於反代 prefix/header/登入驗證
# =========================================================
@require_GET
def whoami(request: HttpRequest) -> JsonResponse:
    data = {
        "ok": True,
        # env / settings
        "ENV_PROXY_PREFIX": os.getenv("PROXY_PREFIX"),
        "ENV_USE_X_FORWARDED_HOST": os.getenv("USE_X_FORWARDED_HOST"),
        "ENV_TRUST_X_FORWARDED_PREFIX": os.getenv("TRUST_X_FORWARDED_PREFIX"),
        "SETTINGS_PROXY_PREFIX": getattr(settings, "PROXY_PREFIX", None),
        "FORCE_SCRIPT_NAME": getattr(settings, "FORCE_SCRIPT_NAME", None),
        "STATIC_URL": getattr(settings, "STATIC_URL", None),
        "MEDIA_URL": getattr(settings, "MEDIA_URL", None),

        # forwarded headers
        "HTTP_X_FORWARDED_HOST": request.META.get("HTTP_X_FORWARDED_HOST", ""),
        "HTTP_X_FORWARDED_PROTO": request.META.get("HTTP_X_FORWARDED_PROTO", ""),
        "HTTP_X_FORWARDED_PREFIX": request.META.get("HTTP_X_FORWARDED_PREFIX", ""),
        "HTTP_X_FORWARDED_FOR": request.META.get("HTTP_X_FORWARDED_FOR", ""),

        # script/path
        "request_script_name": getattr(request, "script_name", ""),
        "META_SCRIPT_NAME": request.META.get("SCRIPT_NAME", ""),
        "META_PATH_INFO": request.META.get("PATH_INFO", ""),
        "request_path": request.path,
        "request_path_info": request.path_info,

        # user (from middleware)
        "login_user": getattr(request, "login_user", "") or "",
        "login_user_name": getattr(request, "login_user_name", "") or "",
        "login_user_org": getattr(request, "login_user_org", "") or "",
        "login_user_org_label": getattr(request, "login_user_org_label", "") or "",
        "login_user_factory_plant": getattr(request, "login_user_factory_plant", "") or "",
        "session_login_user_org": (
            (request.session.get("login_user_org") or "").strip()
            if hasattr(request, "session")
            else ""
        ),

        # raw
        "REMOTE_USER": request.META.get("REMOTE_USER", ""),
        "AUTH_USER": request.META.get("AUTH_USER", ""),

        # oracle debug
        "ora_emp_error": get_last_error(),
    }

    # acl debug
    try:
        from webapps.portal.acl import acl_debug
        data["acl"] = acl_debug(getattr(request, "user", None))
    except Exception as e:
        data["acl"] = {"error": str(e)}

    # ✅ DEBUG 才顯示：確認 .env 是否真的被 load 進 process env
    if getattr(settings, "DEBUG", False):
        data.update({
            "SETTINGS_DEBUG": True,
            "ENV_DEBUG": os.getenv("DEBUG"),
            "ENV_DEV_LOGIN_USER": os.getenv("DEV_LOGIN_USER"),
            "ENV_DEV_LOGIN_NAME": os.getenv("DEV_LOGIN_NAME"),
        })

    return JsonResponse(data, json_dumps_params={"indent": 4, "ensure_ascii": False})

# =========================================================
# helpers
# =========================================================
def _parse_date(s: str) -> Optional[dt.date]:
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def _get_filters(request: HttpRequest) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """
    回傳 (filters_for_queryset, ui_state)
    """
    q = request.GET

    d1 = _parse_date(q.get("date_from", ""))
    d2 = _parse_date(q.get("date_to", ""))
    user_id = (q.get("user_id") or "").strip()
    program_code = (q.get("program_code") or "").strip()

    filters: Dict[str, Any] = {}
    if d1 and d2:
        filters["used_date__range"] = (d1, d2)
    elif d1:
        filters["used_date__gte"] = d1
    elif d2:
        filters["used_date__lte"] = d2

    if user_id:
        filters["user_id__icontains"] = user_id

    if program_code:
        filters["program_code"] = program_code

    ui = {
        "date_from": d1.strftime("%Y-%m-%d") if d1 else "",
        "date_to": d2.strftime("%Y-%m-%d") if d2 else "",
        "user_id": user_id,
        "program_code": program_code,
    }
    return filters, ui


def _today_stats() -> Dict[str, Any]:
    """
    今日統計：
    - 今日使用人數（distinct user_id, 排除空字串）
    - 各功能使用次數
    - 熱門功能排行（Top 10）
    """
    today = timezone.localdate()
    qs_today = PortalUsageLog.objects.filter(used_date=today)

    today_user_count = (
        qs_today.exclude(user_id="")
        .values("user_id")
        .distinct()
        .count()
    )

    program_counts = list(
        qs_today.values("program_code")
        .annotate(cnt=Count("id"))
        .order_by("-cnt")
    )
    top_programs = program_counts[:10]

    return {
        "today": today,
        "today_user_count": today_user_count,
        "program_counts": program_counts,
        "top_programs": top_programs,
    }


def _normalize_plant_label(plant: str) -> str:
    p = (plant or "").strip().upper()
    if p == "MPC":
        return "MPC"
    return p


def _resolve_display_user_name(user_id: str, user_name: str, cache: Dict[str, str]) -> str:
    uid = (user_id or "").strip()
    name0 = (user_name or "").strip()
    if not uid and name0:
        return name0
    if uid in cache:
        return cache[uid]

    # Existing value (except legacy placeholder)
    if name0 and name0 != "使用者":
        if "-" in name0:
            prefix, rest = name0.split("-", 1)
            name0 = f"{_normalize_plant_label(prefix)}-{rest}"
        cache[uid] = name0
        return name0

    # Fallback lookup for old rows with empty/"使用者"
    name = (get_emp_name(uid) or "").strip() if uid else ""
    if name:
        plant = (get_factory_plant_by_id(uid) or "").strip()
        if plant:
            name = f"{_normalize_plant_label(plant)}-{name}"
        cache[uid] = name
        return name

    cache[uid] = ""
    return ""


# =========================================================
# 1) usage page
# =========================================================
@require_node("usage")
def usage_log_page(request: HttpRequest) -> HttpResponse:
    filters, ui = _get_filters(request)

    base_qs = PortalUsageLog.objects.filter(**filters).order_by("-created_at")

    program_codes = list(
        PortalUsageLog.objects.values_list("program_code", flat=True)
        .distinct()
        .order_by("program_code")
    )

    page = int((request.GET.get("page") or "1").strip() or 1)
    page = max(page, 1)
    page_size = 50
    start = (page - 1) * page_size
    end = start + page_size

    total = base_qs.count()
    rows = list(base_qs[start:end])
    name_cache: Dict[str, str] = {}
    for r in rows:
        display_name = _resolve_display_user_name(r.user_id, r.user_name, name_cache)
        setattr(r, "display_user_name", display_name or "（姓名未註記）")

    stats = _today_stats()

    ctx = {
        "ui": ui,
        "rows": rows,
        "total": total,
        "page": page,
        "page_size": page_size,
        "has_prev": page > 1,
        "has_next": end < total,
        "program_codes": program_codes,
        "stats": stats,
    }
    return render(request, "portal/usage.html", ctx)


# =========================================================
# 2) export excel
# - 下載不一定帶 XHR header，故不強制 api=True（避免誤判）
# =========================================================
@require_node("usage")
def usage_log_export_xlsx(request: HttpRequest) -> HttpResponse:
    """
    /portal/usage/export.xlsx (或 /portal/usage/export.xlsx/)
    同查詢條件匯出，並多做 3 張彙總表：
      - Daily（每日）
      - ByUser（每人）
      - ByProgram（每功能 + Top10）
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    filters, _ui = _get_filters(request)
    qs = PortalUsageLog.objects.filter(**filters).order_by("-created_at")

    daily = list(
        qs.values("used_date")
        .annotate(cnt=Count("id"))
        .order_by("used_date")
    )

    by_user = list(
        qs.exclude(user_id="")
        .values("user_id", "user_name")
        .annotate(cnt=Count("id"))
        .order_by("-cnt", "user_id")
    )

    by_program = list(
        qs.values("program_code")
        .annotate(cnt=Count("id"))
        .order_by("-cnt", "program_code")
    )
    top10 = by_program[:10]

    wb = Workbook()

    # ---- Sheet 1: Logs
    ws = wb.active
    ws.title = "Logs"
    ws.append(["created_at", "used_date", "program_code", "user_id", "user_name", "path", "method", "ip"])
    for r in qs.iterator(chunk_size=500):
        ws.append(
            [
                timezone.localtime(r.created_at).strftime("%Y-%m-%d %H:%M:%S"),
                r.used_date.strftime("%Y-%m-%d") if r.used_date else "",
                r.program_code or "",
                r.user_id or "",
                r.user_name or "",
                r.path or "",
                r.method or "",
                r.ip or "",
            ]
        )

    # 自動欄寬（簡版）
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # ---- Sheet 2: Daily
    ws2 = wb.create_sheet("Daily")
    ws2.append(["used_date", "count"])
    for x in daily:
        ws2.append([x["used_date"].strftime("%Y-%m-%d"), x["cnt"]])

    # ---- Sheet 3: ByUser
    ws3 = wb.create_sheet("ByUser")
    ws3.append(["user_id", "user_name", "count"])
    for x in by_user:
        ws3.append([x["user_id"], x.get("user_name") or "", x["cnt"]])

    # ---- Sheet 4: ByProgram
    ws4 = wb.create_sheet("ByProgram")
    ws4.append(["program_code", "count"])
    for x in by_program:
        ws4.append([x["program_code"], x["cnt"]])

    # ---- Sheet 5: Top10
    ws5 = wb.create_sheet("Top10")
    ws5.append(["rank", "program_code", "count"])
    for i, x in enumerate(top10, start=1):
        ws5.append([i, x["program_code"], x["cnt"]])

    # ---- output
    import io

    bio = io.BytesIO()
    wb.save(bio)
    content = bio.getvalue()

    today_tag = timezone.localdate().strftime("%Y%m%d")
    filename = f"portal_usage_{today_tag}.xlsx"

    resp = HttpResponse(
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f"attachment; filename*=UTF-8''{filename}"
    resp["Cache-Control"] = "no-store"
    return resp


@require_node("usage")
def usage_whoami_page(request: HttpRequest) -> HttpResponse:
    """
    Show whoami_json detail for a usage log row.
    URL: /portal/usage/whoami/?id=<log_id>
    """
    log_id_raw = (request.GET.get("id") or "").strip()
    row = None
    err = ""
    pretty_json = ""
    raw_json = ""

    try:
        log_id = int(log_id_raw)
    except Exception:
        log_id = 0

    if log_id <= 0:
        err = "invalid id"
    else:
        row = PortalUsageLog.objects.filter(id=log_id).first()
        if not row:
            err = "log not found"
        else:
            raw_json = (row.whoami_json or "").strip()
            if not raw_json:
                raw_json = "{}"
            try:
                obj = json.loads(raw_json)
                # Hide sensitive/noisy fields in UI (legacy rows may still contain these keys).
                hide_keys = {
                    "aaa",
                    "HTTP_X_FORWARDED_FOR",
                    "HTTP_X_FORWARDED_PROTO",
                    "HTTP_X_FORWARDED_HOST",
                    "HTTP_X_FORWARDED_PREFIX",
                    "HTTP_X_ORIGINAL_FOR",
                    "HTTP_X_CLIENT_IP",
                    "HTTP_FORWARDED",
                    "HTTP_USER_AGENT",
                    "HTTP_REFERER",
                    "session_key",
                    "session_login_user",
                    "session_login_user_name",
                }
                if isinstance(obj, dict):
                    obj = {k: v for k, v in obj.items() if k not in hide_keys}
                pretty_json = json.dumps(obj, ensure_ascii=False, indent=2)
            except Exception:
                pretty_json = raw_json

    ctx = {
        "row": row,
        "error": err,
        "pretty_json": pretty_json,
        "raw_json": raw_json,
        "log_id": log_id_raw,
    }
    return render(request, "portal/usage_whoami.html", ctx)


# =========================================================
# 3) usage user ACL detail
# =========================================================
def _row_get_value(row: Any, key: str, idx: int = 0) -> Any:
    try:
        if hasattr(row, "_mapping"):
            m = row._mapping
            if key in m:
                return m.get(key)
            k2 = key.upper()
            if k2 in m:
                return m.get(k2)
            k3 = key.lower()
            if k3 in m:
                return m.get(k3)
    except Exception:
        pass
    try:
        return row[idx]
    except Exception:
        return None


def _fetch_oracle_acl_groups(user_id: str) -> Tuple[List[Dict[str, str]], str]:
    """
    Return list of {group_name} and error message if any.
    """
    user_id = (user_id or "").strip()
    if not user_id:
        return [], "missing user_id"

    # Strip domain and MPC- prefix for Oracle query safety
    from webapps.portal.middleware import _strip_domain
    user_id = _strip_domain(user_id)
    if user_id.upper().startswith("MPC-"):
        user_id = user_id[4:].strip()

    table = (getattr(settings, "ORA_ACL_TABLE", "") or "").strip()
    user_col = (getattr(settings, "ORA_ACL_USER_COL", "") or "").strip()
    group_col = (getattr(settings, "ORA_ACL_GROUP_COL", "") or "GROUP_NAME").strip()

    if not (table and user_col and group_col):
        return [], "oracle acl settings missing"

    sql = f"SELECT {group_col} AS GROUP_NAME FROM {table} WHERE TRIM({user_col}) = :login_user"
    rows: List[Any] = []
    err = ""
    try:
        rows = db_query_all("oracle", sql, {"login_user": user_id}) or []
    except Exception as e:
        return [], str(e)

    out: List[Dict[str, str]] = []
    for r in rows:
        name = _row_get_value(r, "GROUP_NAME", 0)
        out.append({"group_name": ("" if name is None else str(name).strip())})
    return out, err


def _compute_portal_nodes(user_id: str, groups: List[Dict[str, str]]) -> List[str]:
    acl_map = getattr(settings, "PORTAL_ACL", {}) or {}
    if not acl_map:
        return []

    if not getattr(settings, "PORTAL_ACL_ENABLED", False):
        return sorted(acl_map.keys())

    allow_groups = set()
    for g in groups or []:
        if g.get("group_name"):
            allow_groups.add(g["group_name"])

    allowed = []
    for node, rule in acl_map.items():
        rule_set = {str(x).strip() for x in (rule or []) if str(x).strip()}
        if "PUBLIC" in rule_set:
            allowed.append(node)
            continue
        if "ALL_AUTHENTICATED" in rule_set and user_id:
            allowed.append(node)
            continue
        if allow_groups.intersection(rule_set):
            allowed.append(node)
            continue
    return sorted(allowed)


@require_node("usage")
def usage_user_acl_page(request: HttpRequest) -> HttpResponse:
    user_id = (request.GET.get("user_id") or "").strip()
    user_name = (request.GET.get("user_name") or "").strip()
    if not user_name and user_id:
        user_name = (get_emp_name(user_id) or "").strip()

    groups, err = _fetch_oracle_acl_groups(user_id)
    allowed_nodes = _compute_portal_nodes(user_id, groups)

    ctx = {
        "user_id": user_id,
        "user_name": user_name,
        "groups": groups,
        "allowed_nodes": allowed_nodes,
        "error": err,
    }
    return render(request, "portal/usage_user_acl.html", ctx)
